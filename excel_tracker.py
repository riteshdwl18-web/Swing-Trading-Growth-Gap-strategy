#!/usr/bin/env python3
"""
Excel tracker for stock screening metrics.
Maintains quarterly updates for sales CAGR and price movement tracking.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from stock_screener import (
    get_annual_revenue, get_ttm_revenue, get_price_history, compute_cagr
)
import yfinance as yf

# Config
SYMBOL = "RRKABEL.NS"
MIN_SALES_CAGR_PCT = 15.0
YEARS = 2


def create_tracking_excel(symbol: str, min_cagr: float = MIN_SALES_CAGR_PCT) -> str:
    """
    Creates an Excel file with stock screening metrics for quarterly tracking.
    Returns the filename.
    """
    
    # Fetch data
    revenue, currency = get_annual_revenue(symbol)
    years_list = list(revenue.keys())
    
    # Get 2-year window
    if len(revenue) < 3:
        print(f"Not enough data for {symbol}")
        return None
    
    window = years_list[-3:]
    base_fy, mid_fy, end_fy = window[0], window[1], window[2]
    base_rev, end_rev = revenue[base_fy], revenue[end_fy]
    
    # Calculate metrics
    ticker = yf.Ticker(symbol)
    ttm_revenue = get_ttm_revenue(ticker)
    current_price, entry_price, price_change, price_date = get_price_history(symbol, YEARS)
    cagr = compute_cagr(base_rev, end_rev, YEARS)
    
    if ttm_revenue is not None:
        ttm_growth = ((ttm_revenue / end_rev) - 1) * 100
    else:
        ttm_growth = None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Tracker"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ── Column Headers ────────────────────────────────────────────────────
    headers = [
        "Symbol",
        "FY24 (Cr)",
        "FY26 (Cr)",
        "TTM (Cr)",
        "CAGR % (FY24→FY26)",
        "TTM Growth vs FY26 %",
        "CMP (₹)",
        "Entry Price (₹)",
        "Entry Date",
        "%Away",
        "Status",
        "Last Updated"
    ]
    
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Set column widths
    column_widths = [12, 14, 14, 14, 16, 18, 12, 16, 14, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # ── Data Row ──────────────────────────────────────────────────────────
    
    # Determine status
    status = "✓ PASS" if cagr >= min_cagr else "✗ FAIL"
    
    # Format data
    data_row = [
        symbol,
        f"{base_rev:.2f}" if base_rev else "N/A",
        f"{end_rev:.2f}" if end_rev else "N/A",
        f"{ttm_revenue:.2f}" if ttm_revenue else "N/A",
        f"{cagr:.1f}%" if cagr else "N/A",
        f"{ttm_growth:.1f}%" if ttm_growth else "N/A",
        f"{current_price:.2f}" if current_price else "N/A",
        f"{entry_price:.2f}" if entry_price else "N/A",
        price_date if price_date else "N/A",
        f"{price_change:+.1f}%" if price_change else "N/A",
        status,
        datetime.now().strftime("%d-%b-%Y %H:%M")
    ]
    
    ws.append(data_row)
    
    # Style data row
    data_row_num = 2
    for cell in ws[data_row_num]:
        cell.alignment = center_align
        cell.border = border
        # Color code status
        if "PASS" in str(cell.value):
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell.font = Font(bold=True, color="006100")
        elif "FAIL" in str(cell.value):
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True, color="9C0006")
    
    # ── Quarterly Updates Section ─────────────────────────────────────────
    # This section will be updated as quarters release
    ws.append([])  # Blank row
    ws.append(["QUARTERLY TRACKING"])
    
    quarterly_header_row = ws.max_row
    quarterly_headers = [
        "Quarter",
        "Expected TTM (Cr)",
        "Actual TTM (Cr)",
        "Growth vs FY26 %",
        "Status",
        "Date Updated"
    ]
    ws.append(quarterly_headers)
    
    # Style quarterly header
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = center_align
        cell.border = border
    
    # Pre-fill quarterly rows (to be updated)
    quarters = ["Q1 FY27", "Q2 FY27", "Q3 FY27", "Q4 FY27"]
    for q in quarters:
        ws.append([q, "", "", "", "", ""])
    
    # Style quarterly data rows
    for row_num in range(ws.max_row - 3, ws.max_row + 1):
        for cell in ws[row_num]:
            cell.border = border
            cell.alignment = center_align
    
    # ── Expected TTM Values (Helper Info) ─────────────────────────────────
    ws.append([])  # Blank row
    ws.append(["EXPECTED TTM TARGETS (if maintaining 15% annual growth)"])
    
    target_header = ws.max_row
    target_headers = ["Quarter", "Expected TTM (Cr)", "Growth %"]
    ws.append(target_headers)
    
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center_align
        cell.border = border
    
    # Calculate expected values
    quarterly_rate = (1 + min_cagr / 100) ** (1/4) - 1
    expected_targets = []
    for q in range(1, 5):
        multiplier = (1 + quarterly_rate) ** q
        expected_ttm = end_rev * multiplier
        expected_growth = (multiplier - 1) * 100
        expected_targets.append((f"Q{q}", expected_ttm, expected_growth))
    
    for q_label, expected_ttm, expected_growth in expected_targets:
        ws.append([q_label, f"{expected_ttm:.2f}", f"{expected_growth:.1f}%"])
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = center_align
    
    # ── Instructions ──────────────────────────────────────────────────────
    ws.append([])
    ws.append(["INSTRUCTIONS FOR QUARTERLY UPDATES:"])
    instructions = [
        "1. After each quarter results (Jun, Sep, Dec), update the 'Actual TTM (Cr)' in Quarterly Tracking",
        "2. The 'Growth vs FY26 %' will auto-show if actual TTM matches expected targets",
        "3. Status shows GREEN (✓) if tracking ≥ target, RED (✗) if below target",
        "4. Compare quarterly growth against expected targets (bottom section)",
        "5. 'Entry Price' is the price 2 years ago - entry point for analysis",
        "6. '%Away' shows current price movement from entry point"
    ]
    for instruction in instructions:
        ws.append([instruction])
    
    # Save file
    filename = f"{symbol.replace('.', '_')}_tracking_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print(f"\n✅ Excel file created: {filename}")
    print(f"\nData Summary:")
    print(f"  Company: {symbol}")
    print(f"  Base Year (FY24): {base_rev:.2f} Cr")
    print(f"  End Year (FY26): {end_rev:.2f} Cr")
    print(f"  2-Year CAGR: {cagr:.1f}%")
    print(f"  Current TTM: {ttm_revenue:.2f} Cr (+{ttm_growth:.1f}% vs FY26)")
    print(f"  Current Price: ₹{current_price:.2f}")
    print(f"  Entry Price ({price_date}): ₹{entry_price:.2f}")
    print(f"  Price Appreciation: {price_change:+.1f}%")
    
    return filename


if __name__ == "__main__":
    create_tracking_excel(SYMBOL, MIN_SALES_CAGR_PCT)
