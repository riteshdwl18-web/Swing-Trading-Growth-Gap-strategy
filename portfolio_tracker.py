#!/usr/bin/env python3
"""
Portfolio Tracker - Single consolidated sheet for tracking multiple stocks.
Update quarterly when new results are announced.

USAGE:
    Single stock:   python3 portfolio_tracker.py RRKABEL.NS
    From CSV:       python3 portfolio_tracker.py growth-gap-strategy-mp.csv
    Default (csv):  python3 portfolio_tracker.py
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from stock_screener import (
    get_annual_revenue, get_ttm_revenue, get_price_history, compute_cagr
)
import yfinance as yf
import os
import sys
import csv

# Config
MIN_SALES_CAGR_PCT = 15.0
YEARS = 2
TRACKER_FILE = "Portfolio_Tracker.xlsx"
DEFAULT_CSV = "growth-gap-strategy-mp.csv"


def read_stocks_from_csv(csv_file: str) -> list:
    """
    Read NSE symbols from CSV file.
    Expected columns: Name, BSE Code, NSE Code
    Returns list of symbols with .NS suffix appended.
    """
    stocks = []
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                nse_code = row.get('NSE Code', '').strip()
                if nse_code and nse_code.upper() not in ['', 'NSE CODE']:
                    symbol = f"{nse_code}.NS"
                    stocks.append(symbol)
        
        if stocks:
            print(f"✓ Loaded {len(stocks)} stocks from {csv_file}")
        return stocks
    except FileNotFoundError:
        print(f"  ⚠️  File not found: {csv_file}")
        return []
    except Exception as e:
        print(f"  ⚠️  Error reading {csv_file}: {str(e)}")
        return []


def get_stocks_to_track(input_arg: str = None) -> list:
    """
    Determine which stocks to track based on input.
    Priority:
    1. Command-line argument (single stock or CSV file)
    2. Default CSV file (if exists)
    3. Fall back to single stock
    """
    
    if input_arg:
        # Check if it's a CSV file
        if input_arg.lower().endswith('.csv'):
            stocks = read_stocks_from_csv(input_arg)
            if stocks:
                return stocks
            else:
                print(f"  ⚠️  Could not load stocks from {input_arg}, using default single stock")
        else:
            # Assume it's a single stock symbol
            return [input_arg]
    
    # Try default CSV
    if os.path.exists(DEFAULT_CSV):
        print(f"No input provided. Using default CSV: {DEFAULT_CSV}")
        stocks = read_stocks_from_csv(DEFAULT_CSV)
        if stocks:
            return stocks
    
    # Fall back to single stock
    print("No CSV file found. Using single stock: RRKABEL.NS")
    return ["RRKABEL.NS"]


def get_stock_metrics(symbol: str) -> dict:
    """
    Fetch all metrics for a single stock.
    Returns dict with all required data, or None if insufficient data.
    """
    try:
        # Fetch annual revenue
        revenue, currency = get_annual_revenue(symbol)
        years_list = list(revenue.keys())
        
        if len(revenue) < 3:
            return None
        
        # Get 2-year window
        window = years_list[-3:]
        base_fy, mid_fy, end_fy = window[0], window[1], window[2]
        base_rev = revenue[base_fy]
        end_rev = revenue[end_fy]
        
        # Get current metrics
        ticker = yf.Ticker(symbol)
        ttm_revenue = get_ttm_revenue(ticker)
        current_price, entry_price, price_change, price_date = get_price_history(symbol, YEARS)
        cagr = compute_cagr(base_rev, end_rev, YEARS)
        
        if ttm_revenue is not None:
            ttm_growth = ((ttm_revenue / end_rev) - 1) * 100
        else:
            ttm_growth = 0
        
        status = "✓ PASS" if cagr and cagr >= MIN_SALES_CAGR_PCT else "✗ FAIL"
        
        return {
            "symbol": symbol,
            "base_fy": base_fy,
            "end_fy": end_fy,
            "base_rev": base_rev,
            "end_rev": end_rev,
            "ttm_revenue": ttm_revenue or 0,
            "ttm_growth": ttm_growth,
            "cagr": cagr or 0,
            "current_price": current_price or 0,
            "entry_price": entry_price or 0,
            "price_change": price_change or 0,
            "price_date": price_date or "N/A",
            "status": status,
            "last_updated": datetime.now().strftime("%d-%b-%Y %H:%M")
        }
    except Exception as e:
        print(f"  ⚠️  Error fetching {symbol}: {str(e)}")
        return None


def create_or_update_tracker(stocks_to_track: list = None):
    """
    Create or update the master portfolio tracking file.
    """
    
    if stocks_to_track is None:
        stocks_to_track = ["RRKABEL.NS"]
    
    # Check if file exists
    file_exists = os.path.exists(TRACKER_FILE)
    
    if file_exists:
        wb = load_workbook(TRACKER_FILE)
        ws_master = wb.active
        print(f"✓ Opened existing tracker: {TRACKER_FILE}")
    else:
        wb = Workbook()
        ws_master = wb.active
        ws_master.title = "Master Tracker"
        print(f"✓ Creating new tracker: {TRACKER_FILE}")
    
    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1: MASTER TRACKER (All stocks current status)
    # ════════════════════════════════════════════════════════════════════════
    
    # Clear existing data but keep headers
    if not file_exists:
        # Create headers
        headers = [
            "Symbol",
            "FY24 (Cr)",
            "FY26 (Cr)",
            "TTM (Cr)",
            "CAGR % (FY24→FY26)",
            "TTM Growth vs FY26 %",
            "CMP (₹)",
            "Entry Price (₹)",
            "Entry Date",
            "%Away",
            "Status",
            "Last Updated"
        ]
        ws_master.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in ws_master[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Set column widths
        column_widths = [12, 14, 14, 14, 16, 18, 12, 16, 14, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws_master.column_dimensions[get_column_letter(i)].width = width
    
    # Remove old data rows (keep header)
    ws_master.delete_rows(2, ws_master.max_row)
    
    # Fetch and add data for each stock
    print(f"\nProcessing {len(stocks_to_track)} stocks...")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    passed_count = 0
    failed_count = 0
    
    for i, symbol in enumerate(stocks_to_track, 1):
        print(f"  [{i}/{len(stocks_to_track)}] {symbol}...", end=" ")
        metrics = get_stock_metrics(symbol)
        
        if metrics is None:
            print("⚠️  Skipped")
            continue
        
        # Add data row
        data_row = [
            metrics["symbol"],
            f"{metrics['base_rev']:.2f}",
            f"{metrics['end_rev']:.2f}",
            f"{metrics['ttm_revenue']:.2f}",
            f"{metrics['cagr']:.1f}%",
            f"{metrics['ttm_growth']:.1f}%",
            f"{metrics['current_price']:.2f}",
            f"{metrics['entry_price']:.2f}",
            metrics['price_date'],
            f"{metrics['price_change']:+.1f}%",
            metrics['status'],
            metrics['last_updated']
        ]
        ws_master.append(data_row)
        
        # Style data row
        row_num = ws_master.max_row
        for i, cell in enumerate(ws_master[row_num], 1):
            cell.alignment = center_align
            cell.border = border
            
            # Color code status
            if "PASS" in metrics["status"]:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(bold=True, color="006100")
                passed_count += 1
            elif "FAIL" in metrics["status"]:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True, color="9C0006")
                failed_count += 1
        
        print("✓ Added")
    
    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2: QUARTERLY UPDATES (Template for tracking results)
    # ════════════════════════════════════════════════════════════════════════
    
    if "Quarterly Updates" not in wb.sheetnames:
        ws_quarterly = wb.create_sheet("Quarterly Updates")
        
        # Title
        ws_quarterly['A1'] = "QUARTERLY RESULTS TRACKING"
        ws_quarterly['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_quarterly['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_quarterly.merge_cells('A1:E1')
        
        ws_quarterly['A2'] = "Update this sheet when quarterly results are announced (Jun, Sep, Dec, Mar)"
        ws_quarterly['A2'].font = Font(italic=True, size=10)
        ws_quarterly.merge_cells('A2:E2')
        
        # Headers for quarterly tracking
        quarterly_headers = [
            "Symbol",
            "Quarter",
            "TTM (Cr)",
            "Growth vs Previous FY %",
            "Date Announced"
        ]
        ws_quarterly.append([])  # Blank row
        ws_quarterly.append(quarterly_headers)
        
        # Style quarterly header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws_quarterly[5]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Set column widths
        ws_quarterly.column_dimensions['A'].width = 12
        ws_quarterly.column_dimensions['B'].width = 14
        ws_quarterly.column_dimensions['C'].width = 14
        ws_quarterly.column_dimensions['D'].width = 24
        ws_quarterly.column_dimensions['E'].width = 16
        
        # Pre-fill template rows for each stock
        for symbol in stocks_to_track:
            for quarter in ["Q1", "Q2", "Q3", "Q4"]:
                quarter_label = f"{quarter} FY27"
                ws_quarterly.append([symbol, quarter_label, "", "", ""])
                
                for cell in ws_quarterly[ws_quarterly.max_row]:
                    cell.border = border
                    cell.alignment = center_align
    
    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3: INSTRUCTIONS
    # ════════════════════════════════════════════════════════════════════════
    
    if "Instructions" not in wb.sheetnames:
        ws_info = wb.create_sheet("Instructions")
        
        ws_info['A1'] = "HOW TO USE THIS TRACKER"
        ws_info['A1'].font = Font(bold=True, size=14)
        
        instructions = [
            "",
            "SHEET 1: MASTER TRACKER",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "• Shows CURRENT status of all tracked stocks",
            "• Columns:",
            "  - FY24/FY26: Base and end year revenues (in crores)",
            "  - TTM: Current Trailing Twelve Months revenue",
            "  - CAGR %: 2-year sales growth rate",
            "  - TTM Growth %: Current TTM vs FY26 baseline",
            "  - CMP (₹): Current Market Price",
            "  - Entry Price/Date: Price 2 years ago (entry reference)",
            "  - %Away: Current price movement from entry",
            "  - Status: ✓ PASS (≥15% CAGR) or ✗ FAIL",
            "",
            "SHEET 2: QUARTERLY UPDATES",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "• Update when quarterly results are announced",
            "• Timing: After Jun, Sep, Dec, Mar results",
            "• Columns:",
            "  - Quarter: Q1, Q2, Q3, or Q4",
            "  - TTM (Cr): TTM value announced (from company results)",
            "  - Growth vs Previous FY %: Calculate as ((TTM / Previous FY) - 1) × 100",
            "  - Date Announced: When results were announced",
            "",
            "HOW TO RUN THIS",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "OPTION 1: Single Stock",
            "   cd /home/ritesh-dev/SwingTrading",
            "   ./swing/bin/python portfolio_tracker.py RRKABEL.NS",
            "",
            "OPTION 2: From CSV File (default)",
            "   ./swing/bin/python portfolio_tracker.py growth-gap-strategy-mp.csv",
            "",
            "OPTION 3: Auto-detect (uses CSV if exists, else single stock)",
            "   ./swing/bin/python portfolio_tracker.py",
            "",
            "QUARTERLY UPDATES (Every 3 months):",
            "   • After results announced, update SHEET 2 with new TTM",
            "   • Run: ./swing/bin/python portfolio_tracker.py [same_input]",
            "   • Master Tracker (Sheet 1) auto-updates with latest data",
            "",
            "EXAMPLE: After Jun FY27 results for RRKABEL.NS",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "Sheet 2 entry:",
            "  Symbol: RRKABEL.NS",
            "  Quarter: Q1 FY27",
            "  TTM (Cr): 10,850 (from company results)",
            "  Growth vs FY26: ((10850 / 9587) - 1) × 100 = 13.2%",
            "  Date Announced: 15-Aug-2026",
            "",
            "INTERPRETATION",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "• GREEN (PASS): Stock maintaining ≥15% annual growth → HOLD",
            "• RED (FAIL): Stock below 15% CAGR → REVIEW/EXIT",
            "• TTM Growth %: Track vs expected quarterly progression",
            "• Price %Away: Positive = appreciation, Negative = depreciation"
        ]
        
        for i, instruction in enumerate(instructions, 1):
            ws_info[f'A{i}'] = instruction
            if "━━" in instruction or "SHEET" in instruction or "HOW TO" in instruction or "OPTION" in instruction:
                ws_info[f'A{i}'].font = Font(bold=True, size=11, color="1F4E78")
        
        ws_info.column_dimensions['A'].width = 100
    
    # Save file
    wb.save(TRACKER_FILE)
    print(f"\n✅ Portfolio tracker updated: {TRACKER_FILE}")
    print(f"   • Master Tracker sheet: {passed_count + failed_count} stocks")
    print(f"     - PASS (≥15% CAGR): {passed_count}")
    print(f"     - FAIL (<15% CAGR): {failed_count}")
    print(f"   • Quarterly Updates sheet: Ready for results entry")
    print(f"   • Instructions sheet: Reference guide")
    
    return TRACKER_FILE


if __name__ == "__main__":
    # Parse command-line arguments
    input_arg = sys.argv[1] if len(sys.argv) > 1 else None
    
    # Get stocks to track
    stocks_to_track = get_stocks_to_track(input_arg)
    
    if not stocks_to_track:
        print("⚠️  No stocks to track. Exiting.")
        sys.exit(1)
    
    # Create or update tracker
    create_or_update_tracker(stocks_to_track)
